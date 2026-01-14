import json

import openpyxl
import openpyxl.styles

# 打开已有的excel
wb = openpyxl.load_workbook("result_exp2.xlsm")
sheet_names = wb.sheetnames
sheet = wb["Sheet4"]

with open('data_excel.json', encoding='UTF-8') as file:
    all_data = json.load(file)

for data in all_data:
    acc = "%.3f" %(float(data[2]) / 15)
    sheet.cell(row=data[1], column=data[0], value=acc)

# wb.save("result_exp2.xlsm")  # 注意 excel被手动打开后，保存会失败
# wb.close()

sheet_2 = wb["Sheet3"]

with open('data_weight.json', encoding='UTF-8') as file:
    all_data = json.load(file)

for data in all_data:
    sheet_2.cell(row=data[1], column=data[0], value=data[2])

wb.save("result_exp2.xlsm")  # 注意 excel被手动打开后，保存会失败
wb.close()