import json
import random
import openpyxl

bias = {
    "I" : 1, "E" : 2, "S" : 3, "N" : 4, "F" : 5, "T" : 6, "P" : 7, "J" : 8
}
MODEL = ["OPENAI", "LLAMA", "QWEN"]

with open('data_excel.json', encoding='UTF-8') as file:
    all_data = json.load(file)

wb = openpyxl.load_workbook("result.xlsm")
sheet = wb["Sheet3"]

for data in all_data:
    # model = random.choice(MODEL)
    Round_num = random.randint(0, 4)
    y_post = data[1] + 1
    x_post = (data[0] - 3) * 9 + 3
    model = MODEL[(data[1]+2)%4]
    if len(data) < 4:
        data.append(sheet.cell(row=y_post, column=1).value)
    sheet.cell(row=y_post, column=x_post, value=data[3])

    load = "exp1_" + model + "_test_" + str(Round_num) + "_score.json"
    with open(load, encoding='UTF-8') as file:
        all_score = json.load(file)
    score = all_score[data[3]]
    for key in score:
        sheet.cell(row=y_post, column=x_post+bias[key], value=score[key])

wb.save("result.xlsm")  # 注意 excel被手动打开后，保存会失败
wb.close()
